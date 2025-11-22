from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, current_app, Response
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import os
import json
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models import Doctor, Schedule, User
from app.extensions import db
from app.holiday_utils.holidays import holiday_helper
from functools import wraps

# 创建排班管理蓝图
schedule_bp = Blueprint('schedules', __name__, url_prefix='/schedules')

# 管理员权限装饰器
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not (current_user.is_admin or current_user.is_super_admin):
            flash('需要管理员权限', 'error')
            return redirect(url_for('schedules.index'))
        return f(*args, **kwargs)
    return decorated_function

@schedule_bp.route('/')
def index():
    """排班管理主页 - 显示当月排班表"""
    # 获取选择的月份，默认为当前月 - 限制最小年份为2025
    current_year = datetime.now().year
    month_str = request.args.get('month')
    if month_str:
        year, month = map(int, month_str.split('-'))
        # 确保年份不早于2025
        if year < 2025:
            year = 2025
    else:
        year, month = max(2025, current_year), datetime.now().month

    # 获取该月的第一天和最后一天
    first_day = datetime(year, month, 1)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)

    # 查询该月的所有排班
    schedules = Schedule.query.filter(
        Schedule.date >= first_day.date(),
        Schedule.date <= last_day.date()
    ).order_by(Schedule.date, Schedule.time_range).all()

    # 获取可用医生列表
    available_doctors = Doctor.query.filter_by(status='在职').all()

    # 月份信息
    month_info = {
        'year': year,
        'month': month,
        'days': (last_day - first_day).days + 1
    }

    # 获取当前月份字符串
    current_month = f"{datetime.now().year}-{datetime.now().month:02d}"

    # 获取该月所有日期和星期信息（用于模板显示）
    dates = []
    weekdays = []
    for day in range(1, (last_day - first_day).days + 2):
        current_date = first_day + timedelta(days=day - 1)
        dates.append(current_date.strftime('%Y-%m-%d'))
        weekdays.append(['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日'][current_date.weekday()])

    # 获取年份和月份列表供选择
    years = list(range(2025, 2031))  # 2025-2030年
    months = [
        {'name': '1月', 'value': 1},
        {'name': '2月', 'value': 2},
        {'name': '3月', 'value': 3},
        {'name': '4月', 'value': 4},
        {'name': '5月', 'value': 5},
        {'name': '6月', 'value': 6},
        {'name': '7月', 'value': 7},
        {'name': '8月', 'value': 8},
        {'name': '9月', 'value': 9},
        {'name': '10月', 'value': 10},
        {'name': '11月', 'value': 11},
        {'name': '12月', 'value': 12}
    ]

    # 获取所有医生列表（用于模板显示）
    doctors = Doctor.query.all()

    return render_template('schedules/template.html',
                         schedules=schedules,
                         current_month_str=current_month,
                         years=years,
                         months=months,
                         selected_year=year,
                         selected_month=month,
                         dates=dates,
                         weekdays=weekdays,
                         doctors=doctors,
                         holiday_helper=holiday_helper)

@schedule_bp.route('/generate', methods=['POST'])
@admin_required
def generate_schedule():
    """生成下个月排班表"""
    try:
        target_month = request.form.get('targetMonth')
        use_previous_rules = request.form.get('usePreviousRules') == 'on'

        if not target_month:
            return jsonify({'success': False, 'message': '请选择目标月份'})

        year, month = map(int, target_month.split('-'))
        # 确保年份不早于2025
        if year < 2025:
            return jsonify({'success': False, 'message': '不支持生成2025年以前的排班表，请选择2025年及以后的月份'})
        first_day = datetime(year, month, 1)

        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)

        # 获取该月需要排班的日期（只考虑工作日）
        work_days = []
        current_day = first_day
        while current_day <= last_day:
            # 跳过周末（周六、周日）
            if current_day.weekday() < 5:  # 0-4 表示周一到周五
                work_days.append(current_day)
            current_day += timedelta(days=1)

        # 清理该月现有的排班（避免重复）
        existing_schedules = Schedule.query.filter(
            Schedule.date >= first_day.date(),
            Schedule.date <= last_day.date()
        ).all()
        for schedule in existing_schedules:
            db.session.delete(schedule)

        # 生成排班（这里先使用简单的模板）
        for day in work_days:
            # 每天生成班次（示例：上午班、下午班）
            shifts = [
                {
                    'date': day.date(),
                    'weekday': day.strftime('%A'),
                    'shift': '白班',
                    'time_range': '08:00-16:00',
                    'department': '门诊',
                    'status': 'unassigned'
                },
                {
                    'date': day.date(),
                    'weekday': day.strftime('%A'),
                    'shift': '夜班',
                    'time_range': '16:00-24:00',
                    'department': '急诊',
                    'status': 'unassigned'
                }
            ]

            for shift_data in shifts:
                schedule = Schedule(
                    date=shift_data['date'],
                    weekday=shift_data['weekday'],
                    shift=shift_data['shift'],
                    time_range=shift_data['time_range'],
                    department=shift_data['department'],
                    status=shift_data['status']
                )
                db.session.add(schedule)

        db.session.commit()

        return jsonify({'success': True, 'message': f'成功生成{year}年{month}月排班表，共{len(work_days)}天{len(work_days)*2}个班次'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@schedule_bp.route('/assign_doctor', methods=['POST'])
@admin_required
def assign_doctor():
    """分配医生到排班"""
    try:
        schedule_id = request.form.get('scheduleId')
        doctor_id = request.form.get('doctorSelect')

        if not schedule_id or not doctor_id:
            return jsonify({'success': False, 'message': '请选择排班和医生'})

        schedule = Schedule.query.get(schedule_id)
        doctor = Doctor.query.get(doctor_id)

        if not schedule:
            return jsonify({'success': False, 'message': '排班不存在'})

        if not doctor:
            return jsonify({'success': False, 'message': '医生不存在'})

        # 检查医生是否已有时间冲突的排班
        conflict_schedule = Schedule.query.filter_by(doctor_id=doctor_id).filter(
            Schedule.date == schedule.date,
            Schedule.time_range == schedule.time_range
        ).first()

        if conflict_schedule:
            return jsonify({'success': False, 'message': f'{doctor.name}在该时间段已有排班'})

        # 分配医生
        schedule.doctor_id = doctor_id
        schedule.status = 'assigned'
        db.session.commit()

        return jsonify({'success': True, 'message': f'成功将{doctor.name}分配到{schedule.shift}'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@schedule_bp.route('/<int:schedule_id>')
def view_schedule(schedule_id):
    """查看排班详情"""
    schedule = Schedule.query.get_or_404(schedule_id)
    return render_template('schedules/view.html', schedule=schedule)

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_schedule_file(file_path, file_type):
    """解析排班表文件"""
    # TODO: 需要安装pandas来支持文件解析功能
    return None, "文件解析功能需要安装pandas库，暂时不可用"

@schedule_bp.route('/upload_schedule', methods=['POST'])
@admin_required
def upload_schedule():
    """上传排班表文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})

        file = request.files['file']
        year_month = request.form.get('year_month')
        is_preview = request.form.get('preview') == 'true'

        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})

        if not year_month:
            return jsonify({'success': False, 'message': '请选择年月'})

        # 检查文件类型
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': '不支持的文件类型，请上传Excel或CSV文件'})

        # 解析年月 - 限制最小年份为2025
        try:
            year, month = map(int, year_month.split('-'))
            # 确保年份不早于2025
            if year < 2025:
                return jsonify({'success': False, 'message': '不支持上传2025年以前的排班表，请选择2025年及以后的月份'})
            target_first_day = datetime(year, month, 1).date()
            if month == 12:
                target_last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                target_last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
        except ValueError:
            return jsonify({'success': False, 'message': '年月格式错误'})

        # 保存上传的文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"

        # 确保上传目录存在
        upload_dir = os.path.join(current_app.instance_path, 'uploads', 'schedules')
        os.makedirs(upload_dir, exist_ok=True)

        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)

        # 解析文件
        file_type = filename.rsplit('.', 1)[1].lower()
        records, error = parse_schedule_file(file_path, file_type)

        # 删除临时文件
        try:
            os.remove(file_path)
        except:
            pass

        if error:
            return jsonify({'success': False, 'message': error})

        if not records:
            return jsonify({'success': False, 'message': '文件中没有找到有效的排班数据'})

        # 如果是预览模式，直接返回解析的数据
        if is_preview:
            return jsonify({
                'success': True,
                'data': records,
                'total': len(records)
            })

        # 实际上传处理
        success_count = 0
        failed_count = 0
        errors = []

        try:
            # 清理该月现有的排班（避免重复）
            existing_schedules = Schedule.query.filter(
                Schedule.date >= target_first_day,
                Schedule.date <= target_last_day
            ).all()

            for schedule in existing_schedules:
                db.session.delete(schedule)

            # 获取医生列表用于匹配
            doctors = Doctor.query.filter_by(status='在职').all()
            doctor_dict = {doctor.name: doctor for doctor in doctors}

            # 处理每条记录
            for i, record in enumerate(records, 1):
                try:
                    # 验证日期是否在目标月份范围内
                    record_date = datetime.strptime(record['date'], '%Y-%m-%d').date()
                    if not (target_first_day <= record_date <= target_last_day):
                        failed_count += 1
                        errors.append(f"第{i}行：日期{record['date']}不在选定的月份范围内")
                        continue

                    # 查找医生
                    doctor = None
                    if record['doctor_name']:
                        doctor = doctor_dict.get(record['doctor_name'])
                        if not doctor:
                            failed_count += 1
                            errors.append(f"第{i}行：找不到医生'{record['doctor_name']}'")
                            continue

                    # 创建排班记录
                    schedule = Schedule(
                        doctor_id=doctor.id if doctor else None,
                        date=record_date,
                        weekday=record['weekday'],
                        shift=record['shift'],
                        time_range=record['time_range'],
                        department=record['department'],
                        status='assigned' if doctor else 'unassigned'
                    )

                    db.session.add(schedule)
                    success_count += 1

                except Exception as e:
                    failed_count += 1
                    errors.append(f"第{i}行：{str(e)}")

            db.session.commit()

            return jsonify({
                'success': True,
                'message': f'上传完成！成功处理 {success_count} 条记录，失败 {failed_count} 条',
                'stats': {
                    'total': len(records),
                    'success': success_count,
                    'failed': failed_count
                },
                'errors': errors[:10]  # 只返回前10个错误
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'保存数据时出错：{str(e)}'
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'上传处理失败：{str(e)}'
        })


@schedule_bp.route('/download_template')
def download_template():
    """下载排班表模板CSV文件"""
    try:
        # 创建模板数据
        output = StringIO()
        writer = csv.writer(output)

        # 写入UTF-8 BOM，确保Excel正确显示中文
        output.write('\ufeff')

        # 获取选择的月份参数，默认为当前月 - 限制最小年份为2025
        current_year = datetime.now().year
        month_str = request.args.get('month')
        if month_str:
            try:
                year, month = map(int, month_str.split('-'))
                # 确保年份不早于2025
                if year < 2025:
                    year = 2025
            except ValueError:
                year, month = max(2025, current_year), datetime.now().month
        else:
            year, month = max(2025, current_year), datetime.now().month

        # 计算该月天数
        if month == 12:
            days_in_month = 31
        else:
            days_in_month = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days

        # 星期映射 - 与页面显示保持一致
        weekday_names = ['一', '二', '三', '四', '五', '六', '日']

        # 第一行：日期（第一列显示姓名，模拟斜线表头效果）
        date_row = ['姓名']  # 第一列显示医生姓名标识
        weekday_row = ['']  # 第二行第一列留空，模拟被第一行合并的效果
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            date_row.append(date_str)

            # 计算星期 - 使用真实的星期计算，与页面显示保持一致
            weekday_idx = datetime(year, month, day).weekday()
            weekday_row.append(f"周{weekday_names[weekday_idx]}")  # 使用"周一"格式

        # 写入表头
        writer.writerow(date_row)
        writer.writerow(weekday_row)

        # 获取医生数据
        doctors = Doctor.query.filter_by(status='在职').order_by(Doctor.sequence.asc(), Doctor.id.asc()).all()

        if not doctors:
            # 如果没有医生，写入空表格提示
            empty_row = ['暂无医生数据，请先添加医生']
            # 第一行补充空格到其他列
            for day in range(1, days_in_month + 1):
                empty_row.append('')
            writer.writerow(empty_row)
        else:
            # 使用真实医生数据，但排班内容为空
            for doctor in doctors:
                doctor_row = [doctor.name]
                # 其他列都为空
                for day in range(1, days_in_month + 1):
                    doctor_row.append('')
                writer.writerow(doctor_row)

        # 创建响应
        output.seek(0)
        filename = f'schedule_template_{year}{month:02d}.csv'  # 使用英文文件名避免编码问题
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )

        return response

    except Exception as e:
        current_app.logger.error(f"下载模板失败: {str(e)}")
        flash('下载模板失败', 'error')
        return redirect(url_for('schedules.template'))

@schedule_bp.route('/download_template_excel')
def download_template_excel():
    """下载排班表模板Excel文件"""
    try:
        # 获取选择的月份参数，默认为当前月 - 限制最小年份为2025
        current_year = datetime.now().year
        month_str = request.args.get('month')
        if month_str:
            try:
                year, month = map(int, month_str.split('-'))
                # 确保年份不早于2025
                if year < 2025:
                    year = 2025
            except ValueError:
                year, month = max(2025, current_year), datetime.now().month
        else:
            year, month = max(2025, current_year), datetime.now().month

        # 计算该月天数
        if month == 12:
            days_in_month = 31
        else:
            days_in_month = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days

        # 星期映射 - 与页面显示保持一致
        weekday_names = ['一', '二', '三', '四', '五', '六', '日']

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月排班表"

        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        weekend_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        holiday_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 第一列表头特殊样式（模拟斜线表头）
        first_header_fill = PatternFill(start_color='2E5266', end_color='2E5266', fill_type='solid')

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 姓名列
        for col in range(2, days_in_month + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # 第一行：日期（第一列显示姓名，模拟斜线表头效果）
        date_row = ['姓名']  # 第一列显示医生姓名标识
        weekday_row = ['']  # 第二行第一列留空，模拟被第一行合并的效果

        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            date_row.append(date_str)

            # 计算星期
            weekday_idx = datetime(year, month, day).weekday()
            weekday_row.append(f"周{weekday_names[weekday_idx]}")

        # 写入表头并设置样式
        for col_idx, value in enumerate(date_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.font = header_font
            if col_idx == 1:  # 第一列使用特殊颜色
                cell.fill = first_header_fill
            else:
                cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        for col_idx, value in enumerate(weekday_row, 1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.font = header_font
            if col_idx == 1:  # 第一列使用特殊颜色
                cell.fill = first_header_fill
            else:
                cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # 为周末和节假日列设置特殊背景色
        for day in range(1, days_in_month + 1):
            col_idx = day + 1  # +1 因为第一列是姓名
            weekday_idx = datetime(year, month, day).weekday()

            # 检查是否是节假日
            try:
                from app.models import Holiday
                from datetime import date

                holiday = Holiday.query.filter_by(date=date(year, month, day)).first()
                is_holiday = holiday and holiday.type == 'holiday'
            except Exception:
                # 如果查询失败，默认不是节假日
                is_holiday = False

            # 周末或节假日设置背景色
            if weekday_idx >= 5 or is_holiday:
                for row_idx in [1, 2]:  # 表头两行
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if is_holiday:
                        cell.fill = holiday_fill
                    else:
                        cell.fill = weekend_fill

        # 获取医生数据
        doctors = Doctor.query.filter_by(status='在职').order_by(Doctor.sequence.asc(), Doctor.id.asc()).all()

        if not doctors:
            # 如果没有医生，写入空表格提示（合并所有列）
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=days_in_month + 1)
            empty_cell = ws.cell(row=3, column=1, value="暂无医生数据，请先添加医生")
            empty_cell.font = Font(name='微软雅黑', size=12)
            empty_cell.alignment = center_alignment
            empty_cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        else:
            # 使用真实医生数据，但排班内容为空
            for i, doctor in enumerate(doctors, 3):
                doctor_row = [doctor.name]
                # 其他列都为空
                for day in range(1, days_in_month + 1):
                    doctor_row.append('')

                # 写入医生行并设置样式
                for col_idx, value in enumerate(doctor_row, 1):
                    cell = ws.cell(row=i, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment

                    if col_idx == 1:  # 姓名列
                        cell.font = Font(name='微软雅黑', size=11, bold=True)
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    else:  # 排班列为空
                        cell.value = ''

        # 保存到内存
        excel_output = BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        # 生成文件名（使用ASCII安全的文件名）
        filename = f"schedule_template_{year}-{month:02d}.xlsx"

        # 返回Excel文件
        return Response(
            excel_output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        current_app.logger.error(f"下载Excel模板失败: {str(e)}")
        flash('下载Excel模板失败', 'error')
        return redirect(url_for('schedules.template'))

@schedule_bp.route('/holidays/manage')
def manage_holidays():
    """节假日管理页面"""
    from app.models import Holiday
    import sqlalchemy as sa

    year = request.args.get('year', datetime.now().year, type=int)

    # 生成年份选项 - 限制最小年份为2025
    current_year = datetime.now().year
    min_year = max(2025, current_year)  # 确保最小年份为2025
    # 如果请求的年份小于2025，则使用2025
    if year < 2025:
        year = 2025
    years = list(range(min_year, year + 5))  # 从最小年份开始，往后5年

    # 从数据库获取已保存的节假日数据
    holidays_db = Holiday.query.filter(
        sa.extract('year', Holiday.date) == year
    ).order_by(Holiday.date).all()

    # 转换为模板需要的格式
    holidays = {}
    for holiday in holidays_db:
        date_str = holiday.date.strftime('%Y-%m-%d')
        holidays[date_str] = {
            'name': holiday.name,
            'type': holiday.type,
            'is_system': holiday.is_system,
            'source': 'database'
        }

    return render_template('schedules/holiday_management.html',
                         current_year=year,
                         years=years,
                         holidays=holidays)

@schedule_bp.route('/holidays/add', methods=['POST'])
def add_holiday():
    """添加节假日"""
    from app.models import Holiday
    from datetime import date

    try:
        data = request.get_json()
        date_str = data.get('date')
        end_date_str = data.get('end_date')  # 获取结束日期
        name = data.get('name')
        holiday_type = data.get('type', 'custom')

        if not date_str or not name:
            return jsonify({'success': False, 'message': '日期和名称不能为空'})

        # 检查日期格式
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

            # 如果有结束日期，也检查格式
            if end_date_str:
                end_date_obj = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                # 验证结束日期不能早于开始日期
                if end_date_obj < date_obj:
                    return jsonify({'success': False, 'message': '结束日期不能早于开始日期'})

        except ValueError:
            return jsonify({'success': False, 'message': '日期格式错误，请使用YYYY-MM-DD格式'})

        # 如果没有结束日期，设置为开始日期
        if not end_date_str:
            end_date_str = date_str
            end_date_obj = date_obj

        # 检查范围内的每一天是否已存在节假日
        current_date = date_obj
        while current_date <= end_date_obj:
            current_date_str = current_date.strftime('%Y-%m-%d')
            existing_holiday = Holiday.query.filter_by(date=current_date).first()
            if existing_holiday:
                return jsonify({
                    'success': False,
                    'message': f'日期 {current_date_str} 已存在节假日：{existing_holiday.name}（{existing_holiday.type}），请勿重复添加'
                })
            current_date += timedelta(days=1)

        # 添加节假日到数据库（支持日期范围）
        current_date = date_obj
        added_count = 0
        while current_date <= end_date_obj:
            holiday = Holiday(
                date=current_date,
                name=name,
                type=holiday_type,
                is_system=False  # 用户添加的都不是系统预设
            )
            db.session.add(holiday)
            added_count += 1
            current_date += timedelta(days=1)

        db.session.commit()

        # 清理缓存，确保新添加的节假日能立即生效
        holiday_helper.clear_cache(date_obj.year)

        if added_count == 1:
            message = f'成功添加节假日：{name}（{date_str}）'
        else:
            message = f'成功添加节假日：{name}（{date_str} 至 {end_date_str}，共{added_count}天）'

        return jsonify({'success': True, 'message': message})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'添加失败：{str(e)}'})

@schedule_bp.route('/holidays/remove', methods=['POST'])
def remove_holiday():
    """删除节假日"""
    from app.models import Holiday
    from datetime import date

    try:
        data = request.get_json()
        date_str = data.get('date')

        if not date_str:
            return jsonify({'success': False, 'message': '日期不能为空'})

        # 检查日期格式
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({'success': False, 'message': '日期格式错误，请使用YYYY-MM-DD格式'})

        # 从数据库查找并删除节假日
        holiday = Holiday.query.filter_by(date=date_obj).first()
        if not holiday:
            return jsonify({'success': False, 'message': '该日期没有找到节假日记录'})

        # 删除节假日（允许删除系统预设节假日，因为农历日期可能不完全准确）
        db.session.delete(holiday)
        db.session.commit()

        # 清理缓存，确保删除后能立即生效
        holiday_helper.clear_cache(date_obj.year)

        # 根据是否为系统预设显示不同的提示信息
        if holiday.is_system:
            return jsonify({'success': True, 'message': f'成功删除系统预设节假日：{holiday.name}（{date_str}）'})
        else:
            return jsonify({'success': True, 'message': f'成功删除用户节假日：{holiday.name}（{date_str}）'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'删除失败：{str(e)}'})